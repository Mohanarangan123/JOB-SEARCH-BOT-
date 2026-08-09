"""
XlsxWriter — exports job records to a formatted Excel workbook.

Rules:
  - Uses openpyxl directly (no pandas).
  - One row per unique canonical job — no duplicates.
  - Null/missing values → empty cell (never "null", "None", "N/A").
  - Header row: frozen + bold.
  - Reasonable column widths.
  - Does NOT export: raw HTML, content hash, raw storage path,
    internal version history, full description.

Column order (required):
  Job Title | Company | Location | Workplace Type | Experience
  Salary (raw) | Skills | Posted Date | Employment Type | Source
  Job URL | Application URL | Job ID | Search Run ID | First Seen
  Last Seen | Version | Validation Status | Job Status (optional)
"""
from __future__ import annotations

import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# Column definitions
# ─────────────────────────────────────────────────────────────────────────────

COLUMNS = [
    # (header, width, extractor_key)
    ("Job Title",         30, "title"),
    ("Company",           25, "company"),
    ("Location",          20, "location"),
    ("Workplace Type",    18, "work_mode"),
    ("Experience",        18, "experience"),
    ("Salary (raw)",      20, "salary_raw"),
    ("Skills",            35, "skills"),
    ("Posted Date",       16, "posted_date"),
    ("Employment Type",   18, "employment_type"),
    ("Source",            15, "source"),
    ("Job URL",           50, "job_url"),
    ("Application URL",   40, "apply_url"),
    ("Job ID",            50, "job_id"),
    ("Search Run ID",     22, "run_id"),
    ("First Seen",        22, "first_seen_at"),
    ("Last Seen",         22, "last_seen_at"),
    ("Version",           10, "version"),
    ("Validation Status", 20, "validation_status"),
    ("Job Status",        18, "lifecycle_status"),   # optional but included
]

_HEADER_FILL  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)


# ─────────────────────────────────────────────────────────────────────────────
# Record → row extraction
# ─────────────────────────────────────────────────────────────────────────────

def _clean(value: Any) -> Any:
    """
    Return value as-is if meaningful, else None (→ empty cell).
    Never return the strings 'null', 'None', or 'N/A'.
    """
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.lower() in ("null", "none", "n/a", ""):
            return None
        return stripped
    if isinstance(value, list):
        items = [str(v) for v in value if v is not None]
        return ", ".join(items) if items else None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M UTC")
    return value


def _extract_row(doc: Dict[str, Any]) -> Dict[str, Any]:
    """Extract all export fields from a MongoDB job document."""
    details = doc.get("details") or {}
    source  = doc.get("source") or {}
    reqs    = details.get("requirements") or {}
    comp    = details.get("compensation") or {}
    app     = details.get("application") or {}
    company = details.get("company") or {}
    desc    = details.get("description") or {}

    # Experience: prefer LLM-extracted range, fall back to raw text
    exp_min = reqs.get("experience_years_min")
    exp_max = reqs.get("experience_years_max")
    if exp_min is not None and exp_max is not None:
        experience = f"{exp_min}-{exp_max} years"
    elif exp_min is not None:
        experience = f"{exp_min}+ years"
    else:
        experience = None

    # Skills: required + preferred
    req_skills  = reqs.get("required_skills") or []
    pref_skills = reqs.get("preferred_skills") or []
    all_skills  = req_skills + pref_skills

    # Posted date
    posted_at = details.get("posted_at")

    # Version
    retrieval = doc.get("retrieval") or {}
    versions  = doc.get("version_count")

    return {
        "title":             _clean(details.get("title")),
        "company":           _clean(company.get("name")),
        "location":          _clean(details.get("location")),
        "work_mode":         _clean(details.get("work_mode")),
        "experience":        _clean(experience),
        "salary_raw":        _clean(comp.get("raw_text")),
        "skills":            _clean(all_skills),
        "posted_date":       _clean(posted_at),
        "employment_type":   _clean(details.get("employment_type")),
        "source":            _clean(source.get("source_name")),
        "job_url":           _clean(doc.get("canonical_url")),
        "apply_url":         _clean(app.get("apply_url") or app.get("apply_email")),
        "job_id":            _clean(doc.get("canonical_url")),
        "run_id":            None,    # filled in by caller if available
        "first_seen_at":     _clean(doc.get("first_seen_at")),
        "last_seen_at":      _clean(doc.get("last_seen_at")),
        "version":           _clean(versions),
        "validation_status": _clean((doc.get("validation") or {}).get("status")),
        "lifecycle_status":  _clean(doc.get("lifecycle_status")),
    }


# ─────────────────────────────────────────────────────────────────────────────
# XlsxWriter
# ─────────────────────────────────────────────────────────────────────────────

class XlsxWriter:
    """
    Writes a list of job MongoDB documents to a formatted xlsx workbook.
    """

    def write(
        self,
        jobs: List[Dict[str, Any]],
        output_path: str | Path,
        *,
        run_id: Optional[str] = None,
    ) -> Path:
        """
        Write jobs to an xlsx file.

        Args:
            jobs:        List of job MongoDB documents.
            output_path: Full path for the output file.
            run_id:      Optional search run ID to embed in each row.

        Returns:
            Path of the written file.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Jobs"

        # ── Header row ────────────────────────────────────────────────────
        headers = [col[0] for col in COLUMNS]
        ws.append(headers)

        header_row = ws[1]
        for cell in header_row:
            cell.font      = _HEADER_FONT
            cell.fill      = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN

        ws.freeze_panes = "A2"

        # ── Column widths ────────────────────────────────────────────────
        for col_idx, (_, width, _) in enumerate(COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # ── Data rows (deduplicated by canonical_url) ────────────────────
        seen_urls: set = set()
        rows_written = 0

        for doc in jobs:
            url = doc.get("canonical_url")
            if not url:
                continue
            if url in seen_urls:
                continue  # never export duplicate jobs
            seen_urls.add(url)

            row_data = _extract_row(doc)
            if run_id:
                row_data["run_id"] = run_id

            # Build row in column order
            row = []
            for _, _, key in COLUMNS:
                val = row_data.get(key)
                row.append(val)   # None → empty cell

            ws.append(row)
            rows_written += 1

        # ── Row height for readability ───────────────────────────────────
        for row_num in range(2, rows_written + 2):
            ws.row_dimensions[row_num].height = 16

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))

        logger.info(
            "XlsxWriter: wrote %d rows to %s", rows_written, output_path
        )
        return output_path
