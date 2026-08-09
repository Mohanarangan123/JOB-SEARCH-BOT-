"""
Tests for FastAPI endpoints — no real network, uses mongomock.
"""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict
from unittest.mock import MagicMock

import pytest
from fastapi.testclient import TestClient

try:
    import mongomock
    _MONGOMOCK = True
except ImportError:
    _MONGOMOCK = False


# ─────────────────────────────────────────────────────────────────────────────
# Fixtures
# ─────────────────────────────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def client():
    from main import app
    return TestClient(app)


@pytest.fixture
def db():
    if not _MONGOMOCK:
        pytest.skip("mongomock not installed")
    return mongomock.MongoClient()["test_api"]


@pytest.fixture
def job_repo(db):
    from job_discovery.repositories.job_repository import JobRepository
    return JobRepository(db)


@pytest.fixture
def search_repo(db):
    from job_discovery.repositories.search_repository import SearchRepository
    return SearchRepository(db)


def _sample_job() -> Dict[str, Any]:
    return {
        "canonical_url": "https://naukri.com/job-listings-python-dev-12345678",
        "content_hash": "a" * 64,
        "lifecycle_status": "active",
        "source": {"source_name": "naukri"},
        "details": {
            "title": "Senior Python Developer",
            "location": "Chennai",
            "employment_type": "full-time",
            "work_mode": "hybrid",
            "company": {"name": "TechCorp"},
            "description": {"raw_text": "Python backend role"},
            "requirements": {
                "required_skills": ["Python", "FastAPI"],
                "experience_years_min": 3,
                "experience_years_max": 6,
            },
            "compensation": {"raw_text": "8-14 LPA"},
            "application": {"apply_email": "careers@techcorp.in"},
        },
        "first_seen_at": datetime.now(timezone.utc),
        "last_seen_at":  datetime.now(timezone.utc),
        "last_verified_at": datetime.now(timezone.utc),
        "validation": {"status": "partial", "score": 0.72},
        "version_count": 1,
        "created_at": datetime.now(timezone.utc),
        "updated_at":  datetime.now(timezone.utc),
    }


def _override(app, key, value):
    app.dependency_overrides[key] = lambda: value
    return app


# ─────────────────────────────────────────────────────────────────────────────
# Basic
# ─────────────────────────────────────────────────────────────────────────────

class TestBasicEndpoints:
    def test_health(self, client):
        assert client.get("/health").status_code == 200

    def test_status(self, client):
        r = client.get("/api/status")
        assert r.status_code == 200
        assert r.json()["service"] == "job_discovery"

    def test_openapi(self, client):
        assert client.get("/openapi.json").status_code == 200


# ─────────────────────────────────────────────────────────────────────────────
# POST /api/jobs/search
# ─────────────────────────────────────────────────────────────────────────────

class TestSearchEndpoint:
    def test_trigger_returns_run_id(self, client, db):
        from main import app
        from job_discovery.api.dependencies import get_search_repo
        from job_discovery.repositories.search_repository import SearchRepository

        sr = SearchRepository(db)
        app.dependency_overrides[get_search_repo] = lambda: sr
        try:
            r = client.post("/api/jobs/search", json={"title": "Python Developer"})
            assert r.status_code == 200
            data = r.json()
            assert data["run_id"].startswith("SR_")
        finally:
            app.dependency_overrides.clear()

    def test_empty_body_ok(self, client, db):
        from main import app
        from job_discovery.api.dependencies import get_search_repo
        from job_discovery.repositories.search_repository import SearchRepository

        sr = SearchRepository(db)
        app.dependency_overrides[get_search_repo] = lambda: sr
        try:
            assert client.post("/api/jobs/search", json={}).status_code == 200
        finally:
            app.dependency_overrides.clear()


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/jobs
# ─────────────────────────────────────────────────────────────────────────────

class TestListJobs:
    def _setup(self, app, db):
        from job_discovery.api.dependencies import get_job_repo, get_scorer
        from job_discovery.repositories.job_repository import JobRepository
        from job_discovery.ranking.relevance_scorer import RelevanceScorer
        jr = JobRepository(db)
        sc = RelevanceScorer()
        app.dependency_overrides[get_job_repo] = lambda: jr
        app.dependency_overrides[get_scorer]   = lambda: sc
        return jr

    def test_returns_list(self, client, db):
        from main import app
        self._setup(app, db)
        try:
            r = client.get("/api/jobs")
            assert r.status_code == 200
            assert isinstance(r.json(), list)
        finally:
            app.dependency_overrides.clear()

    def test_returns_job(self, client, db):
        from main import app
        jr = self._setup(app, db)
        jr._jobs.insert_one(_sample_job())
        try:
            r = client.get("/api/jobs?lifecycle_status=active")
            assert r.status_code == 200
            jobs = r.json()
            assert len(jobs) >= 1
            assert jobs[0]["canonical_url"] is not None
        finally:
            app.dependency_overrides.clear()

    def test_sort_newest(self, client, db):
        from main import app
        self._setup(app, db)
        try:
            assert client.get("/api/jobs?sort=newest").status_code == 200
        finally:
            app.dependency_overrides.clear()

    def test_sort_invalid_422(self, client):
        assert client.get("/api/jobs?sort=bad_mode").status_code == 422

    def test_no_raw_html_in_response(self, client, db):
        from main import app
        jr = self._setup(app, db)
        doc = {**_sample_job(), "raw_content_path": "/secret/path"}
        jr._jobs.insert_one(doc)
        try:
            r = client.get("/api/jobs?lifecycle_status=active")
            assert r.status_code == 200
            for job in r.json():
                assert "raw_content_path" not in job
                assert "page_text" not in job
        finally:
            app.dependency_overrides.clear()

    def test_relevance_sort_with_criteria(self, client, db):
        from main import app
        self._setup(app, db)
        try:
            r = client.get("/api/jobs?sort=relevance&title=Python+Developer&location=Chennai")
            assert r.status_code == 200
        finally:
            app.dependency_overrides.clear()


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/jobs/detail
# ─────────────────────────────────────────────────────────────────────────────

class TestGetJob:
    def test_get_existing_job(self, client, db):
        from main import app
        from job_discovery.api.dependencies import get_job_repo
        from job_discovery.repositories.job_repository import JobRepository

        jr = JobRepository(db)
        jr._jobs.insert_one(_sample_job())
        app.dependency_overrides[get_job_repo] = lambda: jr
        try:
            url = _sample_job()["canonical_url"]
            r = client.get(f"/api/jobs/detail?canonical_url={url}")
            assert r.status_code == 200
            assert r.json()["title"] == "Senior Python Developer"
        finally:
            app.dependency_overrides.clear()

    def test_get_missing_job_404(self, client, db):
        from main import app
        from job_discovery.api.dependencies import get_job_repo
        from job_discovery.repositories.job_repository import JobRepository

        jr = JobRepository(db)
        app.dependency_overrides[get_job_repo] = lambda: jr
        try:
            r = client.get("/api/jobs/detail?canonical_url=https://fake.com/job/99")
            assert r.status_code == 404
        finally:
            app.dependency_overrides.clear()

    def test_missing_canonical_url_422(self, client):
        assert client.get("/api/jobs/detail").status_code == 422


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/jobs/versions
# ─────────────────────────────────────────────────────────────────────────────

class TestJobVersions:
    def test_versions_endpoint(self, client, db):
        from main import app
        from job_discovery.api.dependencies import get_job_repo
        from job_discovery.repositories.job_repository import JobRepository

        jr = JobRepository(db)
        url = "https://naukri.com/job-listings-versions-test-99"
        jr.save_version(url, {
            "version_number": 1,
            "content_hash": "a" * 64,
            "retrieved_at": datetime.now(timezone.utc),
        })
        app.dependency_overrides[get_job_repo] = lambda: jr
        try:
            r = client.get(f"/api/jobs/versions?canonical_url={url}")
            assert r.status_code == 200
            data = r.json()
            assert isinstance(data, list)
            assert data[0]["version_number"] == 1
        finally:
            app.dependency_overrides.clear()

    def test_missing_canonical_url_422(self, client):
        assert client.get("/api/jobs/versions").status_code == 422


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/jobs/changes
# ─────────────────────────────────────────────────────────────────────────────

class TestJobChanges:
    def test_changes_endpoint(self, client, db):
        from main import app
        from job_discovery.api.dependencies import get_job_repo
        from job_discovery.repositories.job_repository import JobRepository

        jr = JobRepository(db)
        url = "https://naukri.com/job-listings-changes-test-99"
        jr.save_version(url, {
            "version_number": 2,
            "content_hash": "b" * 64,
            "retrieved_at": datetime.now(timezone.utc),
            "changes": {
                "changed_fields": ["salary_raw"],
                "details": [{"field": "salary_raw", "before": "8 LPA", "after": "12 LPA"}],
            },
        })
        app.dependency_overrides[get_job_repo] = lambda: jr
        try:
            r = client.get(f"/api/jobs/changes?canonical_url={url}")
            assert r.status_code == 200
            changes = r.json()
            assert isinstance(changes, list)
            assert changes[0]["field_name"] == "salary_raw"
            assert changes[0]["before"] == "8 LPA"
            assert changes[0]["after"] == "12 LPA"
        finally:
            app.dependency_overrides.clear()


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/search-runs/{run_id}
# ─────────────────────────────────────────────────────────────────────────────

class TestSearchRunEndpoint:
    def test_get_existing_run(self, client, db):
        from main import app
        from job_discovery.api.dependencies import get_search_repo
        from job_discovery.repositories.search_repository import SearchRepository
        from job_discovery.models.search import SearchRun, RunStatus

        sr = SearchRepository(db)
        sr.runs.insert(SearchRun(run_id="SR_20260808_TEST1", status=RunStatus.completed))
        app.dependency_overrides[get_search_repo] = lambda: sr
        try:
            r = client.get("/api/search-runs/SR_20260808_TEST1")
            assert r.status_code == 200
            assert r.json()["run_id"] == "SR_20260808_TEST1"
            assert r.json()["status"] == "completed"
        finally:
            app.dependency_overrides.clear()

    def test_missing_run_404(self, client, db):
        from main import app
        from job_discovery.api.dependencies import get_search_repo
        from job_discovery.repositories.search_repository import SearchRepository

        sr = SearchRepository(db)
        app.dependency_overrides[get_search_repo] = lambda: sr
        try:
            assert client.get("/api/search-runs/SR_NOTEXIST").status_code == 404
        finally:
            app.dependency_overrides.clear()


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/sources/health
# ─────────────────────────────────────────────────────────────────────────────

class TestSourcesHealth:
    def test_returns_list(self, client):
        r = client.get("/api/sources/health")
        assert r.status_code == 200
        assert isinstance(r.json(), list)


# ─────────────────────────────────────────────────────────────────────────────
# POST /api/jobs/export
# ─────────────────────────────────────────────────────────────────────────────

class TestExportEndpoint:
    def test_export_response(self, client, db):
        from main import app
        from job_discovery.api.dependencies import get_job_repo
        from job_discovery.repositories.job_repository import JobRepository

        jr = JobRepository(db)
        app.dependency_overrides[get_job_repo] = lambda: jr
        try:
            r = client.post("/api/jobs/export")
            assert r.status_code == 200
            data = r.json()
            assert "message" in data
            assert "job_count" in data
        finally:
            app.dependency_overrides.clear()

    def test_xlsx_export_file_schema(self, tmp_path):
        from openpyxl import load_workbook
        from job_discovery.export.xlsx_writer import export_jobs_to_xlsx, EXPORT_COLUMNS

        doc = {
            "canonical_url": "https://naukri.com/jobs/python-dev-123",
            "external_job_id": "NKR-123",
            "content_hash": "a" * 64,
            "lifecycle_status": "active",
            "search_run_id": "SR_TEST",
            "first_seen_at": datetime.now(timezone.utc),
            "last_seen_at": datetime.now(timezone.utc),
            "version_count": 1,
            "validation": {"status": "valid"},
            "source": {"source_name": "naukri"},
            "details": {
                "title": "Python Developer",
                "location": "Chennai",
                "work_mode": "hybrid",
                "employment_type": "full-time",
                "posted_at": datetime.now(timezone.utc),
                "company": {"name": "Example Co"},
                "requirements": {
                    "required_skills": ["Python", "FastAPI"],
                    "experience_years_min": 0,
                    "experience_years_max": 2,
                },
                "compensation": {"raw_text": "12-18 LPA"},
                "application": {"apply_url": "https://example.com/apply"},
            },
        }

        out = export_jobs_to_xlsx([doc], output_path=str(tmp_path / "jobs.xlsx"), overwrite_target=False)
        assert Path(out).exists()
        wb = load_workbook(out)
        ws = wb.active
        assert ws.title == "Jobs"
        assert [c.value for c in ws[1]] == EXPORT_COLUMNS
        assert ws.max_row == 2
        assert ws["A2"].value == "Python Developer"
        assert ws["B2"].value == "Example Co"
        assert ws["K2"].value == "https://naukri.com/jobs/python-dev-123"
        assert ws["L2"].value == "https://example.com/apply"
        assert ws["M2"].value == "NKR-123"
        assert ws["N2"].value == "SR_TEST"
        assert ws.freeze_panes == "A2"
        assert ws["A1"].font.bold
